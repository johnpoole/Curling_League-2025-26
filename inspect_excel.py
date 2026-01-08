import openpyxl

excel_path = r"c:\Users\jdpoo\Documents\GitHub\Curling_League 2025-26\2025 2026 Curling Roster All Leagues (1).xlsx"
wb = openpyxl.load_workbook(excel_path)

print("Sheets in workbook:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"Sheet: {sheet_name}")
    print('='*80)
    
    # Print first 20 rows to see the structure
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if row_idx > 20:
            break
        # Only show rows with content
        if any(cell is not None for cell in row):
            print(f"Row {row_idx}: {row}")
