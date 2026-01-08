import openpyxl
import json
import re

def extract_all_rosters(excel_path):
    """Extract all team rosters from the Excel file"""
    wb = openpyxl.load_workbook(excel_path)
    all_rosters = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"Processing sheet: {sheet_name}")
        print('='*80)
        
        # Get all data from the sheet
        data = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data.append(row)
        
        # Find the header row (look for "Team Name" or similar)
        header_row_idx = None
        team_name_col = None
        player_cols = []
        
        for idx, row in enumerate(data):
            for col_idx, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    cell_lower = str(cell).lower().strip()
                    if 'team' in cell_lower and 'name' in cell_lower:
                        header_row_idx = idx
                        team_name_col = col_idx
                        # Find player columns
                        for i, header_cell in enumerate(row):
                            if header_cell and (str(header_cell).strip().startswith('Player') or 
                                               str(header_cell).strip() in ['Skip', 'Vice', 'Second', 'Lead']):
                                player_cols.append(i)
                        break
            if header_row_idx is not None:
                break
        
        if header_row_idx is None or team_name_col is None:
            print(f"Could not find header row in {sheet_name}")
            continue
        
        print(f"Found header at row {header_row_idx}, team column {team_name_col}")
        print(f"Player columns: {player_cols}")
        
        # Extract team rosters
        for row_idx in range(header_row_idx + 1, len(data)):
            row = data[row_idx]
            
            if team_name_col < len(row):
                team_name = row[team_name_col]
                
                if team_name and str(team_name).strip() and str(team_name).strip() != 'None':
                    team_name = str(team_name).strip()
                    
                    # Extract players
                    players = []
                    for player_col in player_cols:
                        if player_col < len(row) and row[player_col]:
                            player_name = str(row[player_col]).strip()
                            if player_name and player_name != 'None':
                                players.append(player_name)
                    
                    if players:
                        # Determine league from sheet name
                        league = sheet_name
                        if 'monday' in sheet_name.lower():
                            league = "Monday Night"
                        elif 'tuesday' in sheet_name.lower():
                            league = "Tuesday Night"
                        
                        # Store roster with league info
                        key = f"{team_name}|{league}"
                        all_rosters[key] = {
                            "team": team_name,
                            "league": league,
                            "players": players
                        }
                        print(f"  {team_name} ({league}): {', '.join(players)}")
    
    return all_rosters

# Extract all rosters
print("\n" + "="*80)
print("EXTRACTING ALL TEAM ROSTERS FROM EXCEL")
print("="*80)

excel_path = r"c:\Users\jdpoo\Documents\GitHub\Curling_League 2025-26\2025 2026 Curling Roster All Leagues (1).xlsx"
all_rosters = extract_all_rosters(excel_path)

# Load existing Poole data
with open('poole_team_data.json', 'r') as f:
    poole_data = json.load(f)

# Update Poole's roster with correct info
poole_monday_key = "Poole|Monday Night"
poole_tuesday_key = "Poole|Tuesday Night"

# Update the rosters section
updated_rosters = []
if poole_monday_key in all_rosters:
    updated_rosters.append({
        "league": "Monday Night",
        "players": all_rosters[poole_monday_key]["players"]
    })

if poole_tuesday_key in all_rosters:
    updated_rosters.append({
        "league": "Tuesday Night",
        "players": all_rosters[poole_tuesday_key]["players"]
    })

poole_data["rosters"] = updated_rosters

# Add all rosters to the data structure
poole_data["all_team_rosters"] = {}
for key, roster_info in all_rosters.items():
    team = roster_info["team"]
    league = roster_info["league"]
    
    if team not in poole_data["all_team_rosters"]:
        poole_data["all_team_rosters"][team] = {}
    
    poole_data["all_team_rosters"][team][league] = roster_info["players"]

# Save updated data
with open('poole_team_data.json', 'w', encoding='utf-8') as f:
    json.dump(poole_data, f, indent=2, ensure_ascii=False)

print("\n" + "="*80)
print("SUMMARY")
print("="*80)
print(f"Total teams extracted: {len(set(r['team'] for r in all_rosters.values()))}")
print(f"Total roster entries: {len(all_rosters)}")
print("\nPoole Team Rosters:")
for roster in updated_rosters:
    print(f"\n{roster['league']}:")
    for player in roster['players']:
        print(f"  - {player}")

print("\n✓ Updated poole_team_data.json with all team rosters")
print("\nYou can now look up any opponent's roster from the database!")
