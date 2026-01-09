import openpyxl
import json
import re

def extract_all_rosters_monday(sheet):
    """Extract rosters from Monday Men's sheet"""
    rosters = {}
    
    # The format appears to be: team numbers in rows, then 4 rows of players per team
    # Teams are in columns with 2 columns per team (name, email)
    
    # Read all data
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    # Process the sheet - teams are laid out in blocks
    # Look for team numbers (integers) which mark the start of a team block
    i = 0
    while i < len(data):
        row = data[i]
        
        # Find team numbers in this row
        team_positions = []
        for col_idx, cell in enumerate(row):
            if cell is not None and isinstance(cell, int) and 1 <= cell <= 50:
                team_positions.append((cell, col_idx))
        
        if team_positions:
            # Next 4-5 rows contain player names for these teams
            for team_num, col_idx in team_positions:
                players = []
                team_name = None
                
                # Check if team number includes name (like "8 - Plaid Lads")
                if isinstance(row[col_idx], str) and '-' in str(row[col_idx]):
                    parts = str(row[col_idx]).split('-', 1)
                    team_name = parts[1].strip()
                
                # Look for all players (variable number per team)
                for offset in range(1, 10):  # Check more rows to get all players
                    if i + offset < len(data):
                        next_row = data[i + offset]
                        if col_idx < len(next_row):
                            player = next_row[col_idx]
                            if player and isinstance(player, str) and '@' not in player:
                                player = player.strip()
                                if player and not re.match(r'^\d+$', player) and player not in ['None', '']:
                                    # Check if it's a new team number (stop collecting)
                                    if isinstance(player, str) and re.match(r'^\d+\s*-?\s*', player):
                                        break
                                    players.append(player)
                
                # Try to identify team name from first player's last name if not found
                if not team_name and players:
                    # Use last word of first player's name
                    team_name = players[0].split()[-1]
                
                if players and team_name:
                    rosters[team_name] = {
                        "team": team_name,
                        "league": "Monday Night",
                        "players": players  # Take all players found
                    }
        
        i += 1
    
    return rosters

def extract_all_rosters_tuesday(sheet):
    """Extract rosters from Tuesday Men's sheet"""
    rosters = {}
    
    # Similar structure to Monday
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    # Process the sheet
    i = 0
    while i < len(data):
        row = data[i]
        
        # Find team info - look for "1 - Team Name" pattern or team numbers
        team_positions = []
        for col_idx, cell in enumerate(row):
            if cell is not None:
                # Check for "number - name" pattern
                if isinstance(cell, str) and re.match(r'^\d+\s*-\s*.+', str(cell)):
                    match = re.match(r'^(\d+)\s*-\s*(.+)', str(cell))
                    if match:
                        team_num = int(match.group(1))
                        team_name = match.group(2).strip()
                        team_positions.append((team_name, col_idx))
                # Check for just a number
                elif isinstance(cell, int) and 1 <= cell <= 50:
                    team_positions.append((None, col_idx))
        
        if team_positions:
            for team_name, col_idx in team_positions:
                players = []
                
                # Look for all players (variable number per team)
                for offset in range(1, 10):  # Check more rows to get all players
                    if i + offset < len(data):
                        next_row = data[i + offset]
                        if col_idx < len(next_row):
                            player = next_row[col_idx]
                            if player and isinstance(player, str) and '@' not in player:
                                player = player.strip()
                                if player and not re.match(r'^\d+$', player) and player not in ['None', '']:
                                    # Skip if it's another team header
                                    if not re.match(r'^\d+\s*-\s*.+', player):
                                        players.append(player)
                                    else:
                                        break  # Stop at next team
                
                # If no team name, use first player's last name
                if not team_name and players:
                    team_name = players[0].split()[-1]
                
                if players and team_name:
                    rosters[team_name] = {
                        "team": team_name,
                        "league": "Tuesday Night",
                        "players": players  # Take all players found
                    }
        
        i += 1
    
    # Also check the pool list on the right side (columns around 10-11)
    # Look for team names in that area
    for i, row in enumerate(data):
        if len(row) > 10:
            # Check for team names in the pool list
            if row[10] and isinstance(row[10], str) and not re.match(r'^\d+$', row[10]):
                team_name = row[10].strip()
                if team_name and team_name not in ['A', 'B', 'Pool', 'None', '']:
                    if team_name not in rosters:
                        # This team doesn't have players yet, try to find them
                        rosters[team_name] = {
                            "team": team_name,
                            "league": "Tuesday Night",
                            "players": []
                        }
    
    return rosters

# Main execution
excel_path = r"c:\Users\jdpoo\Documents\GitHub\Curling_League 2025-26\2025 2026 Curling Roster All Leagues (1).xlsx"
wb = openpyxl.load_workbook(excel_path)

all_rosters = {}

# Process Monday
print("\n" + "="*80)
print("MONDAY NIGHT")
print("="*80)
monday_sheet = wb["Monday Men's"]
monday_rosters = extract_all_rosters_monday(monday_sheet)
all_rosters.update(monday_rosters)

for team, info in sorted(monday_rosters.items()):
    print(f"\n{team}:")
    for player in info['players']:
        print(f"  - {player}")

# Process Tuesday
print("\n" + "="*80)
print("TUESDAY NIGHT")
print("="*80)
tuesday_sheet = wb["Tuesday Men's"]
tuesday_rosters = extract_all_rosters_tuesday(tuesday_sheet)

# Update with Tuesday rosters
for team, info in tuesday_rosters.items():
    key = f"{team}|Tuesday"
    all_rosters[key] = info

for team, info in sorted(tuesday_rosters.items()):
    print(f"\n{team}:")
    for player in info['players']:
        print(f"  - {player}")

# Manually add correct Poole rosters since we know them
poole_monday = {
    "team": "Poole",
    "league": "Monday Night",
    "players": [
        "John Poole",
        "Jason Grelowski",
        "Bryan Jensen",
        "Phillip LaFlair"
    ]
}

poole_tuesday = {
    "team": "Poole",
    "league": "Tuesday Night",
    "players": [
        "John Poole",
        "Jason Grelowski",
        "Thad Snethun",
        "Glen Phelps"
    ]
}

# Override Poole entries
all_rosters["Poole"] = poole_monday
all_rosters["Poole|Tuesday"] = poole_tuesday

# Load existing schedule data
with open('poole_team_data.json', 'r') as f:
    poole_data = json.load(f)

# Update rosters
poole_data["rosters"] = [poole_monday, poole_tuesday]

# Add all team rosters organized by team name
poole_data["all_team_rosters"] = {}
for key, roster_info in all_rosters.items():
    team = roster_info["team"]
    league = roster_info["league"]
    
    if team not in poole_data["all_team_rosters"]:
        poole_data["all_team_rosters"][team] = {}
    
    poole_data["all_team_rosters"][team][league] = roster_info["players"]

# Save
with open('poole_team_data.json', 'w', encoding='utf-8') as f:
    json.dump(poole_data, f, indent=2, ensure_ascii=False)

print("\n" + "="*80)
print("SUMMARY")
print("="*80)
print(f"Total teams extracted: {len(all_rosters)}")
print("\n✓ Updated poole_team_data.json")
print("\nYou can now see opponent rosters for all teams!")
