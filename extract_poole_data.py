import pdfplumber
import openpyxl
import pandas as pd
import json
import re
from datetime import datetime

def extract_poole_from_pdf(pdf_path, league_name):
    """Extract all Poole team games from a PDF schedule"""
    games = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            
            # Print the text to help debug
            if "Poole" in text or "poole" in text.lower():
                print(f"\n=== Found Poole in {league_name} - Page {page.page_number} ===")
                print(text)
                print("=" * 80)
            
            # Split into lines and search for Poole
            lines = text.split('\n')
            for i, line in enumerate(lines):
                if 'poole' in line.lower():
                    # Extract game information from this line and surrounding lines
                    # The format may vary, so we'll need to parse carefully
                    games.append({
                        'line': line,
                        'context': '\n'.join(lines[max(0, i-2):min(len(lines), i+3)]),
                        'page': page.page_number
                    })
    
    return games

def extract_roster_from_excel(excel_path):
    """Extract Poole team roster from Excel file"""
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path)
        
        # Try to find Poole in all sheets
        poole_roster = []
        league_info = None
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\nSearching in sheet: {sheet_name}")
            
            # Search for Poole in the sheet
            for row in sheet.iter_rows(values_only=True):
                row_text = ' '.join([str(cell) for cell in row if cell is not None])
                if 'poole' in row_text.lower():
                    print(f"Found Poole: {row}")
                    if not league_info:
                        league_info = sheet_name
        
        # Also try with pandas
        for sheet_name in wb.sheetnames:
            try:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                print(f"\n=== DataFrame for {sheet_name} ===")
                print(df.head())
                
                # Search for Poole in any column
                for col in df.columns:
                    poole_rows = df[df[col].astype(str).str.contains('poole', case=False, na=False)]
                    if not poole_rows.empty:
                        print(f"\nFound Poole in column '{col}':")
                        print(poole_rows)
                        poole_roster.extend(poole_rows.values.tolist())
            except Exception as e:
                print(f"Error reading sheet {sheet_name} with pandas: {e}")
        
        return poole_roster, league_info
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return [], None

# Main execution
print("=" * 80)
print("EXTRACTING DATA FOR TEAM POOLE")
print("=" * 80)

# Extract from Monday schedule
print("\n\n### MONDAY NIGHT LEAGUE ###")
monday_games = extract_poole_from_pdf(
    r"c:\Users\jdpoo\Documents\GitHub\Curling_League 2025-26\Monday Night Mens 2025 2026 28 1st Round - 28 teams.pdf",
    "Monday Night"
)

# Extract from Tuesday schedule
print("\n\n### TUESDAY NIGHT LEAGUE ###")
tuesday_games = extract_poole_from_pdf(
    r"c:\Users\jdpoo\Documents\GitHub\Curling_League 2025-26\Tuesday Night Mens 2025 2026 1st Round 26 teams.pdf",
    "Tuesday Night"
)

# Extract roster
print("\n\n### ROSTER DATA ###")
roster_data, league_info = extract_roster_from_excel(
    r"c:\Users\jdpoo\Documents\GitHub\Curling_League 2025-26\2025 2026 Curling Roster All Leagues (1).xlsx"
)

# Prepare results
results = {
    "team_name": "Poole",
    "leagues": [],
    "monday_raw_data": monday_games,
    "tuesday_raw_data": tuesday_games,
    "roster_raw_data": roster_data,
    "league_info": league_info
}

if monday_games:
    results["leagues"].append("Monday Night")
if tuesday_games:
    results["leagues"].append("Tuesday Night")

# Print summary
print("\n\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"Team: Poole")
print(f"Leagues: {', '.join(results['leagues']) if results['leagues'] else 'None found'}")
print(f"Monday games found: {len(monday_games)}")
print(f"Tuesday games found: {len(tuesday_games)}")
print(f"Roster entries found: {len(roster_data)}")

# Save to JSON
with open('poole_raw_data.json', 'w') as f:
    json.dump(results, f, indent=2, default=str)

print("\nRaw data saved to poole_raw_data.json")
print("Please review the output to help parse the specific format.")
